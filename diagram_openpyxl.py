from openpyxl import load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.colors import ColorChoice
from openpyxl.utils import get_column_letter
import tkinter.messagebox as messagebox


def diagram(book):
    """
    Построение диаграмм в Excel с использованием openpyxl вместо win32com.
    
    Примечание: openpyxl имеет ограниченные возможности форматирования диаграмм
    по сравнению с win32com. Некоторые функции (раскраска отдельных столбцов,
    вторичные оси для комбинированных диаграмм) могут работать иначе.
    """
    global diagramma_coor, score, Gauss, marks, score_interval, category, start_col
    global diagramma_coor_marks, category_marks, format_diapazone, script_task
    global format_diapazone_x, diagramma_coor_bias, table_massive
    
    try:
        # Загружаем книгу Excel
        wb = load_workbook(book)
        sheet = wb.active
        
        # Скрываем все столбцы (сброс)
        for col in sheet.columns:
            col_letter = get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].hidden = False
        
        # Цвета для оценок (в формате RGB)
        color_map = {
            '2': 'FF7675',   # Красный
            '3': 'FAB1A0',   # Оранжевый
            '4': '55EFC4',   # Зеленый
            '5': '74B9FF'    # Синий
        }
        
        # ==================== ДИАГРАММА 1: Распределение по баллам ====================
        chart1 = BarChart()
        chart1.type = "col"
        chart1.style = 10
        chart1.title = "Сравнение распределение баллов с нормальным"
        chart1.y_axis.title = "доля участников, вероятность"
        chart1.x_axis.title = "Балл"
        chart1.legend = None
        
        # Данные: предполагаем, что таблица For_diagrams существует
        # Нужно найти диапазоны для [Доля], [Балл], [Гаусс]
        # В openpyxl работаем с ссылками на ячейки
        
        # Создаем ссылку на данные для основной серии (Доля)
        data_ref1 = Reference(sheet, min_col=sheet.tables['For_diagrams'].columns['Доля'].min,
                              min_row=sheet.tables['For_diagrams'].ref.min_row + 1,
                              max_row=sheet.tables['For_diagrams'].ref.max_row)
        cats_ref1 = Reference(sheet, min_col=sheet.tables['For_diagrams'].columns['Балл'].min,
                              min_row=sheet.tables['For_diagrams'].ref.min_row + 1,
                              max_row=sheet.tables['For_diagrams'].ref.max_row)
        
        series1 = chart1.add_series(data_ref1)
        chart1.set_categories(cats_ref1)
        
        # Настройка размеров (~пиксели -> единицы openpyxl)
        chart1.height = 14  # ~350 пикселей
        chart1.width = 25   # ~600 пикселей
        
        # Добавление серии Гаусса как линии
        gauss_series = LineChart()
        gauss_data = Reference(sheet, min_col=sheet.tables['For_diagrams'].columns['Гаусс'].min,
                               min_row=sheet.tables['For_diagrams'].ref.min_row + 1,
                               max_row=sheet.tables['For_diagrams'].ref.max_row)
        gauss_series.add_series(gauss_data)
        gauss_series.style = 13  # Синяя линия
        
        # Размещаем первую диаграмму
        sheet.add_chart(chart1, diagramma_coor)
        
        # ==================== ДИАГРАММА 2: Распределение по отметкам ====================
        chart2 = BarChart()
        chart2.type = "col"
        chart2.style = 10
        chart2.title = "Распределение по отметкам ВПР"
        chart2.y_axis.title = "Количество"
        chart2.x_axis.title = "Отметка"
        chart2.legend = None
        
        # Данные из таблицы Marks
        data_ref2 = Reference(sheet, min_col=sheet.tables['Marks'].columns['Количество'].min,
                              min_row=sheet.tables['Marks'].ref.min_row + 1,
                              max_row=sheet.tables['Marks'].ref.max_row)
        cats_ref2 = Reference(sheet, min_col=sheet.tables['Marks'].columns['Оценка'].min,
                              min_row=sheet.tables['Marks'].ref.min_row + 1,
                              max_row=sheet.tables['Marks'].ref.max_row)
        
        series2 = chart2.add_series(data_ref2)
        chart2.set_categories(cats_ref2)
        
        # Настройка подписей данных
        chart2.dataLabels = DataLabelList()
        chart2.dataLabels.showVal = True
        chart2.dataLabels.showCatName = False
        chart2.dataLabels.showSerName = False
        chart2.dataLabels.position = 't'  # сверху
        
        chart2.height = 12  # ~300 пикселей
        chart2.width = 12   # ~300 пикселей
        
        sheet.add_chart(chart2, diagramma_coor_marks)
        
        # ==================== ДИАГРАММА 3: Несоответствие отметок ====================
        chart3 = BarChart()
        chart3.type = "col"
        chart3.style = 10
        chart3.title = "Несоответствие отметок"
        chart3.y_axis.title = "%"
        chart3.x_axis.title = "Категория"
        chart3.legend = None
        
        # Данные из таблицы Bias_marks
        data_ref3 = Reference(sheet, min_col=sheet.tables['Bias_marks'].columns['Процент'].min,
                              min_row=sheet.tables['Bias_marks'].ref.min_row + 1,
                              max_row=sheet.tables['Bias_marks'].ref.max_row)
        cats_ref3 = Reference(sheet, min_col=sheet.tables['Bias_marks'].columns['Критерий'].min,
                              min_row=sheet.tables['Bias_marks'].ref.min_row + 1,
                              max_row=sheet.tables['Bias_marks'].ref.max_row)
        
        series3 = chart3.add_series(data_ref3)
        chart3.set_categories(cats_ref3)
        
        # Настройка подписей данных
        chart3.dataLabels = DataLabelList()
        chart3.dataLabels.showVal = True
        chart3.dataLabels.showCatName = False
        chart3.dataLabels.showSerName = False
        chart3.dataLabels.position = 't'  # сверху
        
        chart3.height = 12  # ~300 пикселей
        chart3.width = 15   # ~370 пикселей
        
        sheet.add_chart(chart3, diagramma_coor_bias)
        
        # ==================== ДИАГРАММА 4: % выполнения заданий ====================
        chart4 = BarChart()
        chart4.type = "col"
        chart4.style = 10
        chart4.title = "% выполнения заданий"
        chart4.y_axis.title = "%"
        chart4.x_axis.title = "Задания"
        
        chart4.height = 16  # ~400 пикселей
        chart4.width = 30   # ~750 пикселей
        
        # Ограничение оси Y до 100%
        chart4.y_axis.max = 100.0
        chart4.y_axis.min = 0.0
        
        # Убираем сетку
        chart4.y_axis.majorGridlines = None
        chart4.x_axis.majorGridlines = None
        
        # Данные из диапазона format_diapazone
        # Предполагаем, что format_diapazone - это строковый диапазон типа "C5:D20"
        data_ref4 = Reference(sheet, range_string=format_diapazone)
        cats_ref4 = Reference(sheet, range_string=script_task)
        
        series4 = chart4.add_series(data_ref4)
        series4.title = "% выполнение задания"
        chart4.set_categories(cats_ref4)
        
        # Вторая серия (Не приступали к заданию)
        data_ref5 = Reference(sheet, range_string=format_diapazone_x)
        series5 = chart4.add_series(data_ref5)
        series5.title = "% не приступивших к заданию"
        
        # Настройка подписей данных с форматом процентов
        chart4.dataLabels = DataLabelList()
        chart4.dataLabels.showVal = True
        chart4.dataLabels.numberFormat = '0%'
        
        # Включаем легенду
        chart4.legend = None  # По умолчанию
        
        sheet.add_chart(chart4, diagramma_coor_task)
        
        # Сохраняем книгу
        wb.save(book)
        wb.close()
        
        print("Диаграммы успешно созданы с использованием openpyxl")
        
    except Exception as e:
        messagebox.showerror("", f"Ошибка при создании диаграмм: {e}")
        print(f"Ошибка при создании диаграмм: {e}")
